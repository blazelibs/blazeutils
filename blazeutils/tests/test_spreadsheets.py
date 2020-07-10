from __future__ import absolute_import, unicode_literals

import datetime as dt
import decimal

import pytest

from blazeutils.spreadsheets import (
    Reader,
    WriterX,
    XlwtHelper,
    http_headers,
    workbook_to_reader,
    xlsx_to_reader,
    xlsx_to_strio,
)
from blazeutils.testing import emits_deprecation


class TestWorkbookToReader(object):

    def test_xlwt_to_reader(self):
        import xlwt

        write_wb = xlwt.Workbook()
        ws = write_wb.add_sheet('Foo')
        ws.write(0, 0, 'bar')

        wb = workbook_to_reader(write_wb)
        sh = wb.sheet_by_name('Foo')
        assert sh.cell_value(rowx=0, colx=0) == 'bar'


class TestXlsxToReader(object):

    def test_xlsx_to_reader(self):
        import xlsxwriter

        write_wb = xlsxwriter.Workbook()
        ws = write_wb.add_worksheet('Foo')
        ws.write(0, 0, 'bar')

        wb = xlsx_to_reader(write_wb)
        sh = wb.sheet_by_name('Foo')
        assert sh.cell_value(rowx=0, colx=0) == 'bar'


class TestWriter(object):

    @emits_deprecation('XlwtHelper has been renamed to Writer')
    def test_xlwt_helper_deprecation(self):
        XlwtHelper()


class TestWriterX:
    def test_unknown_type(self):
        with pytest.raises(TypeError):
            WriterX({'bad': 'format'})


class TestOpenpyxlWriter:
    def test_awrite(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        writer = WriterX(sheet)
        writer.awrite('Pooh')
        writer.awrite('Bar', nextrow=True)
        writer.awrite('Baz')
        assert sheet.cell(row=1, column=1).value == 'Pooh'
        assert sheet.cell(row=1, column=2).value == 'Bar'
        assert sheet.cell(row=2, column=1).value == 'Baz'

    def test_awrite_style(self):
        import openpyxl
        from openpyxl.styles import Font, NamedStyle, PatternFill

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        writer = WriterX(sheet)
        style = NamedStyle(name='test')
        font = Font(bold=True, size=20, color='ff0000')
        fill = PatternFill(start_color='ff0000ff', end_color='ff0000ff', fill_type='solid')
        style.font = font
        style.fill = fill
        writer.awrite('Bar', style=style)
        cell = sheet['A1']
        assert cell.value == 'Bar'
        assert cell.font == font
        assert cell.fill == fill

    def test_mwrite(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        writer = WriterX(sheet)
        row1 = ['Pooh', 'Bar', 'Baz', 'ab', 'cd', 'ef']
        row2 = ['gh', 'ij', 'kl']
        writer.mwrite(*row1[:3])
        writer.mwrite(*row1[3:], nextrow=True)
        writer.mwrite(*row2)
        for colnum in range(1, len(row1) + 1):
            assert sheet.cell(row=1, column=colnum).value == row1[colnum-1]
        for colnum in range(1, len(row2) + 1):
            assert sheet.cell(row=2, column=colnum).value == row2[colnum-1]

    def test_mwrite_style(self):
        import openpyxl
        from openpyxl.styles import Font, NamedStyle, PatternFill

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        writer = WriterX(sheet)
        style = NamedStyle(name='test')
        font = Font(bold=True, size=20, color='ff0000')
        fill = PatternFill(start_color='ff0000ff', end_color='ff0000ff', fill_type='solid')
        style.font = font
        style.fill = fill
        row = ['Pooh', 'Bar', 'Baz']
        writer.mwrite(*row, style=style)
        for colnum in range(1, len(row) + 1):
            cell = sheet.cell(row=1, column=colnum)
            assert cell.value == row[colnum-1]
            assert cell.font == font
            assert cell.fill == fill


class TestXlsxwriterWriter:
    def test_awrite(self):
        import openpyxl
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        writer = WriterX(sheet)
        writer.awrite('Pooh')
        writer.awrite('Bar', nextrow=True)
        writer.awrite('Baz')

        read_workbook = openpyxl.load_workbook(xlsx_to_strio(workbook))
        read_sheet = read_workbook[read_workbook.sheetnames[0]]
        assert read_sheet['A1'].value == 'Pooh'
        assert read_sheet['B1'].value == 'Bar'
        assert read_sheet['A2'].value == 'Baz'

    def test_awrite_style(self):
        import openpyxl
        import xlsxwriter
        from blazeutils.spreadsheets import WriterX

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        writer = WriterX(sheet)
        format = workbook.add_format({
            'bold': True, 'font_color': 'red', 'font_size': 20, 'bg_color': 'blue',
        })
        writer.awrite('Bar', style=format)

        read_workbook = openpyxl.load_workbook(xlsx_to_strio(workbook))
        read_sheet = read_workbook[read_workbook.sheetnames[0]]
        cell = read_sheet['A1']
        assert cell.value == 'Bar'
        assert cell.font.color.value == 'FFFF0000'
        assert cell.font.bold
        assert cell.font.size == 20
        assert cell.fill.fgColor.value == 'FF0000FF'

    def test_mwrite(self):
        import openpyxl
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        writer = WriterX(sheet)
        row1 = ['Pooh', 'Bar', 'Baz', 'ab', 'cd', 'ef']
        row2 = ['gh', 'ij', 'kl']
        writer.mwrite(*row1[:3])
        writer.mwrite(*row1[3:], nextrow=True)
        writer.mwrite(*row2)

        read_workbook = openpyxl.load_workbook(xlsx_to_strio(workbook))
        read_sheet = read_workbook[read_workbook.sheetnames[0]]
        for colnum in range(1, len(row1) + 1):
            assert read_sheet.cell(row=1, column=colnum).value == row1[colnum-1]
        for colnum in range(1, len(row2) + 1):
            assert read_sheet.cell(row=2, column=colnum).value == row2[colnum-1]

    def test_mwrite_style(self):
        import openpyxl
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        writer = WriterX(sheet)
        format = workbook.add_format({
            'bold': True, 'font_color': 'red', 'font_size': 20, 'bg_color': 'blue',
        })
        writer = WriterX(sheet)
        row = ['Pooh', 'Bar', 'Baz']
        writer.mwrite(*row, style=format)

        read_workbook = openpyxl.load_workbook(xlsx_to_strio(workbook))
        read_sheet = read_workbook[read_workbook.sheetnames[0]]
        for colnum in range(1, len(row) + 1):
            cell = read_sheet.cell(row=1, column=colnum)
            assert cell.value == row[colnum-1]
            assert cell.font.color.value == 'FFFF0000'
            assert cell.font.bold
            assert cell.font.size == 20
            assert cell.fill.fgColor.value == 'FF0000FF'


class TestReader:
    def test_unknown_type(self):
        with pytest.raises(TypeError):
            Reader({'bad': 'format'})

    def test_from_xlwt(self):
        import xlwt

        write_wb = xlwt.Workbook()
        ws = write_wb.add_sheet('Foo')
        ws.write(0, 0, 'bar')

        reader = Reader.from_xlwt(write_wb)
        assert reader.cell_value() == 'bar'

    def test_from_xlsx(self):
        import xlsxwriter

        write_wb = xlsxwriter.Workbook()
        ws = write_wb.add_worksheet('Foo')
        ws.write('A1', 'bar')

        reader = Reader.from_xlsx(write_wb)
        assert reader.cell_value() == 'bar'


class TestReaderOpenpyxl:
    def test_cell_value(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet['A1'] = 'Pooh'
        sheet['B1'] = 'Bar'
        reader = Reader(workbook)
        assert reader.cell_value() == 'Pooh'
        assert reader.cell_value() == 'Bar'
        assert reader.cell_value() is None

    def test_cell_date(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet['A1'] = dt.date(2020, 1, 1)
        sheet['B1'] = dt.datetime(2019, 12, 31, 3, 5)
        sheet['C1'] = 'Bla'
        reader = Reader(workbook)
        assert reader.cell_date() == dt.date(2020, 1, 1)
        assert reader.cell_date() == dt.date(2019, 12, 31)
        with pytest.raises(TypeError) as e:
            reader.cell_date()
            assert str(e) == 'Bla: not a date'

    def test_cell_datetime(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet['A1'] = dt.date(2020, 1, 1)
        sheet['B1'] = dt.datetime(2019, 12, 31, 3, 5)
        sheet['C1'] = 'Bla'
        reader = Reader(workbook)
        assert reader.cell_datetime() == dt.datetime(2020, 1, 1, 0, 0)
        assert reader.cell_datetime() == dt.datetime(2019, 12, 31, 3, 5)
        with pytest.raises(TypeError) as e:
            reader.cell_date()
            assert str(e) == 'Bla: not a datetime'

    def test_cell_decimal(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet['A1'] = '1234.56'
        sheet['B1'] = 'Bla'
        reader = Reader(workbook)
        assert reader.cell_decimal() == decimal.Decimal('1234.56')
        with pytest.raises(decimal.InvalidOperation):
            reader.cell_decimal()
        with pytest.raises(TypeError):
            reader.cell_decimal()

    def test_next_row(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet['A1'] = 'A1-text'
        sheet['A2'] = 'A2-text'
        sheet['B2'] = 'B2-text'
        reader = Reader(workbook)
        assert reader.cell_value() == 'A1-text'
        reader.next_row()
        assert reader.cell_value() == 'A2-text'
        assert reader.cell_value() == 'B2-text'


class TestReaderXlrd:
    def test_cell_value(self):
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        sheet.write(0, 0, 'Pooh')
        sheet.write(0, 1, 'Bar')
        reader = Reader.from_xlsx(workbook)
        assert reader.cell_value() == 'Pooh'
        assert reader.cell_value() == 'Bar'

    def test_cell_date(self):
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        sheet.write_datetime('A1', dt.date(2020, 1, 1))
        sheet.write_datetime('B1', dt.datetime(2019, 12, 31, 3, 5))
        sheet.write('C1', 'Bla')
        reader = Reader.from_xlsx(workbook)
        assert reader.cell_date() == dt.date(2020, 1, 1)
        assert reader.cell_date() == dt.date(2019, 12, 31)
        with pytest.raises(TypeError) as e:
            reader.cell_date()
            assert str(e) == 'Bla: not a date'

    def test_cell_datetime(self):
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        sheet.write_datetime('A1', dt.date(2020, 1, 1))
        sheet.write_datetime('B1', dt.datetime(2019, 12, 31, 3, 5))
        sheet.write('C1', 'Bla')
        reader = Reader.from_xlsx(workbook)
        assert reader.cell_datetime() == dt.datetime(2020, 1, 1, 0, 0)
        assert reader.cell_datetime() == dt.datetime(2019, 12, 31, 3, 5)
        with pytest.raises(TypeError) as e:
            reader.cell_date()
            assert str(e) == 'Bla: not a datetime'

    def test_cell_decimal(self):
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        sheet.write('A1', '1234.56')
        sheet.write('B1', 'Bla')
        reader = Reader.from_xlsx(workbook)
        assert reader.cell_decimal() == decimal.Decimal('1234.56')
        with pytest.raises(decimal.InvalidOperation):
            reader.cell_decimal()

    def test_next_row(self):
        import xlsxwriter

        workbook = xlsxwriter.Workbook()
        sheet = workbook.add_worksheet('Foo')
        sheet.write('A1', 'A1-text')
        sheet.write('A2', 'A2-text')
        sheet.write('B2', 'B2-text')
        reader = Reader.from_xlsx(workbook)
        assert reader.cell_value() == 'A1-text'
        reader.next_row()
        assert reader.cell_value() == 'A2-text'
        assert reader.cell_value() == 'B2-text'


class TestHttpHeaders(object):

    def test_xls_filename(self):
        expect = {
            'Content-Type': 'application/vnd.ms-excel',
            'Content-Disposition': 'attachment; filename=foo.xls'
        }
        assert http_headers('foo.xls', randomize=False), expect

    def test_xlsx_filename(self):
        expect = {
            'Content-Type': 'application/vnd.openxmlformats-officedocument'
                            '.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename=foo.xlsx'
        }
        assert http_headers('foo.xlsx', randomize=False) == expect

    def test_randomize(self):
        content_dispo = http_headers('foo.xlsx')['Content-Disposition']
        _, filename = content_dispo.split('=')
        intpart = filename.replace('foo-', '').replace('.xlsx', '')
        assert int(intpart) >= 1000000
